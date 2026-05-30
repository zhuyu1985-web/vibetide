#!/usr/bin/env python3
"""
解析 /Users/zhuyu/Downloads/副本2025年线下生态宣传活动统计表(1).xlsx
输出 /tmp/activities-2025.json
"""
import json
from openpyxl import load_workbook
from datetime import datetime, timedelta

XLSX = "/Users/zhuyu/Downloads/副本2025年线下生态宣传活动统计表(1).xlsx"
OUT = "/tmp/activities-2025.json"

# xlsx 里的区县名 → DB 里 cq_districts 的标准名
NAME_MAP = {
    "西部科学城重庆高新区": "科学城重庆高新区",
    # 其他无需映射
}

ACTIVITY_THEMES = ["六五环境日", "815全国生态日", "志愿服务活动", "环保设施向公众开放", "美丽重庆六进活动"]

wb = load_workbook(XLSX, data_only=True)
ws = wb["Sheet1"]

# 第一区:5-43 行原始数据 (39 区县)
districts = []
for r in range(5, 44):
    seq = ws.cell(row=r, column=2).value
    name = ws.cell(row=r, column=3).value
    if not name:
        continue
    name = NAME_MAP.get(str(name).strip(), str(name).strip())
    counts = [ws.cell(row=r, column=4 + i).value or 0 for i in range(5)]
    total = ws.cell(row=r, column=9).value or sum(counts)
    last_date = ws.cell(row=r, column=10).value
    first_date = ws.cell(row=r, column=11).value
    # excel 日期序号 → 实际日期(以 1899-12-30 为基准)
    if isinstance(last_date, (int, float)) and isinstance(first_date, (int, float)):
        last_dt = datetime(1899, 12, 30) + timedelta(days=int(last_date))
        first_dt = datetime(1899, 12, 30) + timedelta(days=int(first_date))
        days = (last_dt - first_dt).days + 1
        first_iso = first_dt.date().isoformat()
        last_iso = last_dt.date().isoformat()
    else:
        days = None
        first_iso = None
        last_iso = None
    districts.append({
        "seq": int(seq) if seq else None,
        "district": name,
        "themes": dict(zip(ACTIVITY_THEMES, counts)),
        "total": int(total),
        "first_date": first_iso,
        "last_date": last_iso,
        "span_days": days,
        "freq": (int(total) / days) if (days and days > 0) else None,
    })

print(f"✓ 解析得 {len(districts)} 个区县")
for d in districts[:3]:
    print(f"  {d['district']}: total={d['total']}, days={d['span_days']}, freq={d['freq']}")
print(f"  ...")
for d in districts[-3:]:
    print(f"  {d['district']}: total={d['total']}, days={d['span_days']}, freq={d['freq']}")

# 用 DB 中的 39 区县列表校验
DB_DISTRICTS = [
    "万州区", "万盛经开区", "两江新区", "丰都县", "九龙坡区", "云阳县", "北碚区", "南岸区",
    "南川区", "合川区", "垫江县", "城口县", "大渡口区", "大足区", "奉节县", "巫山县", "巫溪县",
    "巴南区", "开州区", "彭水县", "忠县", "梁平区", "武隆区", "永川区", "江津区", "沙坪坝区",
    "涪陵区", "渝中区", "潼南区", "璧山区", "石柱县", "秀山县", "科学城重庆高新区", "綦江区",
    "荣昌区", "酉阳县", "铜梁区", "长寿区", "黔江区",
]
xlsx_names = {d["district"] for d in districts}
db_set = set(DB_DISTRICTS)
miss_in_xlsx = db_set - xlsx_names
miss_in_db = xlsx_names - db_set
if miss_in_xlsx:
    print(f"\n⚠ xlsx 缺失 {len(miss_in_xlsx)} 个: {miss_in_xlsx}")
if miss_in_db:
    print(f"\n⚠ xlsx 多出 {len(miss_in_db)} 个: {miss_in_db}")
if not miss_in_xlsx and not miss_in_db:
    print(f"\n✓ 39 区县完全对齐")

with open(OUT, "w") as f:
    json.dump(districts, f, ensure_ascii=False, indent=2)
print(f"\n✓ 写: {OUT}")
