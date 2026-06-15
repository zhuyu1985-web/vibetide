#!/usr/bin/env python3
"""
基于 /tmp/ranking-v5-2025-from-db.json 生成 ranking-v5-2025.xlsx

10 个 sheet:
  0-总览     综合排名 39 行 (与 docx 表 2-1 一致)
  1-中央     一级指标:中央媒体 - 二级原始 + 二级区间化 + 一级合计
  2-行业     一级指标:行业媒体
  3-市级     一级指标:市级媒体
  4-区县     一级指标:区县媒体
  5-公众     一级指标:公众行为引导 (全 80 占位)
  6-原始数据  39 区县 × 4 tier × 3 子指标 = 12 列原始值
  7-区间化数据 39 区县 × 4 tier × 3 子指标 = 12 列 [65,95] 值
  8-权重    所有权重 + 公式说明
  9-统计    统计量 + 梯队分布
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

JSON_PATH = "/tmp/ranking-v5-2025-from-db.json"
OUT_PATH = "/Users/zhuyu/Developer/chinamcloud/vibetide/docs/ranking-v5-2025.xlsx"

with open(JSON_PATH) as f:
    data = json.load(f)

ranked = data["ranked"]
raw = data["raw"]
scaled = data["scaled"]
topics = data["topics"]
stats = data["stats"]
weights = data["weights"]

# 39 个区县按综合得分降序
districts = [r["name"] for r in ranked]

wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
TIER_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

def style_body(ws, start_row=2):
    for ri in range(start_row, ws.max_row + 1):
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ri % 2 == 0:
                cell.fill = ALT_FILL

def fit_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 0: 综合排名 =====
ws = wb.create_sheet("0-综合排名")
ws.append(["排名", "区县", "中央媒体 (45%)", "行业媒体 (25%)", "市级媒体 (15%)", "区县媒体 (8%)", "公众行为 (7%)", "综合得分"])
for r in ranked:
    ws.append([
        r["rank"], r["name"],
        round(r["central"], 2), round(r["industry"], 2),
        round(r["municipal"], 2), round(r["district"], 2),
        round(r["public"], 2), round(r["composite"], 2),
    ])
style_header_row(ws)
style_body(ws)
fit_widths(ws, [6, 14, 16, 16, 16, 16, 14, 12])

# 加底部统计
ws.append([])
ws.append(["最高", "", "", "", "", "", "", round(stats["max"], 2)])
ws.append(["最低", "", "", "", "", "", "", round(stats["min"], 2)])
ws.append(["分差", "", "", "", "", "", "", round(stats["span"], 2)])
ws.append(["平均", "", "", "", "", "", "", round(stats["mean"], 2)])
ws.append(["中位数", "", "", "", "", "", "", round(stats["median"], 2)])
ws.append(["标准差", "", "", "", "", "", "", round(stats["stdev"], 2)])

# ===== Sheet 1-4: 4 个一级指标 (含媒体数据) =====
TIER_LABELS = {
    "central": ("1-中央媒体传播指数", "中央媒体"),
    "industry": ("2-行业媒体传播指数", "行业媒体"),
    "municipal": ("3-市级媒体传播指数", "市级媒体"),
    "district": ("4-区县媒体传播指数", "区县媒体"),
}
for tier_key, (sheet_name, label) in TIER_LABELS.items():
    ws = wb.create_sheet(sheet_name)
    ws.append([
        "排名", "区县",
        "报道数量 (原始)", "主题丰富度 F (原始)", "传播速度 (原始)",
        "数量得分", "丰富度得分", "速度得分",
        f"{label}指数",
    ])
    # 按本一级指数降序排
    rows_sorted = sorted(districts, key=lambda d: -[r for r in ranked if r["name"] == d][0][tier_key])
    for i, name in enumerate(rows_sorted, 1):
        raw_t = raw[name][tier_key]
        scaled_t = scaled[name][tier_key]
        final = [r for r in ranked if r["name"] == name][0][tier_key]
        ws.append([
            i, name,
            raw_t["count"], round(raw_t["richness"], 4), round(raw_t["freq"], 4),
            round(scaled_t["count"], 2), round(scaled_t["richness"], 2), round(scaled_t["freq"], 2),
            round(final, 2),
        ])
    style_header_row(ws)
    style_body(ws)
    fit_widths(ws, [6, 14, 14, 18, 14, 12, 14, 12, 14])

# ===== Sheet 5: 公众行为引导 =====
ws = wb.create_sheet("5-公众行为引导指数")
ws.append(["排名", "区县", "活动数量", "活动主题丰富度", "活动速度", "公众行为指数", "备注"])
for i, name in enumerate(districts, 1):
    ws.append([i, name, "-", "-", "-", 80, "2025 年度数据待重庆市生态环境局协调各区县上报"])
style_header_row(ws)
style_body(ws)
fit_widths(ws, [6, 14, 12, 16, 12, 16, 50])

# ===== Sheet 6: 原始值矩阵 =====
ws = wb.create_sheet("6-原始数据矩阵")
ws.append([
    "区县",
    "央-数量", "央-丰富度F", "央-速度",
    "业-数量", "业-丰富度F", "业-速度",
    "市-数量", "市-丰富度F", "市-速度",
    "区-数量", "区-丰富度F", "区-速度",
])
for name in districts:
    row = [name]
    for tier in ["central", "industry", "municipal", "district"]:
        r = raw[name][tier]
        row.extend([r["count"], round(r["richness"], 4), round(r["freq"], 4)])
    ws.append(row)
style_header_row(ws)
style_body(ws)
fit_widths(ws, [14] + [12] * 12)

# ===== Sheet 7: 区间化数据矩阵 =====
ws = wb.create_sheet("7-区间化数据矩阵")
ws.append([
    "区县",
    "央-数量", "央-丰富度", "央-速度",
    "业-数量", "业-丰富度", "业-速度",
    "市-数量", "市-丰富度", "市-速度",
    "区-数量", "区-丰富度", "区-速度",
])
for name in districts:
    row = [name]
    for tier in ["central", "industry", "municipal", "district"]:
        r = scaled[name][tier]
        row.extend([round(r["count"], 2), round(r["richness"], 2), round(r["freq"], 2)])
    ws.append(row)
style_header_row(ws)
style_body(ws)
fit_widths(ws, [14] + [12] * 12)

# ===== Sheet 8: 权重 & 公式 =====
ws = wb.create_sheet("8-权重与公式")
ws.append(["项目", "细节"])
rows = [
    ("一级权重 中央", "45%"),
    ("一级权重 行业", "25%"),
    ("一级权重 市级", "15%"),
    ("一级权重 区县", "8%"),
    ("一级权重 公众", "7%"),
    ("", ""),
    ("二级权重 报道数量/活动数量", "40%"),
    ("二级权重 主题丰富度", "30%"),
    ("二级权重 传播速度", "30%"),
    ("", ""),
    ("区间化区间", "[65, 95]"),
    ("主题维度 N", "16 个 (媒体报道)"),
    ("活动主题维度", "5 个 (公众活动)"),
    ("", ""),
    ("丰富度公式", "F = 1 / Σ |p_t − 1/N|, p_t 为某主题占该 (district, tier) 报道总量的比例"),
    ("速度公式", "freq = 报道总数 / 有发布的不同日期数 (按 YYYY-MM-DD 计)"),
    ("综合公式", "score = 中央×0.45 + 行业×0.25 + 市级×0.15 + 区县×0.08 + 公众×0.07"),
    ("一级指标公式", "tier_score = 数量得分×0.40 + 丰富度得分×0.30 + 速度得分×0.30"),
    ("", ""),
    ("tier → outlet 映射 中央", "media_outlet_dictionary.outlet_tier = 'central' (12 家)"),
    ("tier → outlet 映射 行业", "outlet_tier = 'industry' (32 家)"),
    ("tier → outlet 映射 市级", "outlet_tier = 'provincial_municipal' (32 家)"),
    ("tier → outlet 映射 区县", "outlet_tier ∈ ('district_media', 'government_self_media') (84 家)"),
    ("tier → outlet 映射 公众", "全 39 区县固定 80 分占位 (2025 年线下活动数据待补)"),
    ("", ""),
    ("数据窗口", "2025-01-01 ≤ published_at < 2026-01-01"),
    ("数据规模", "61,371 篇 (有 outlet 的 2025 年 collected_items)"),
    ("annotation 来源", "research_collected_item_districts / research_collected_item_topics"),
]
for k, v in rows:
    ws.append([k, v])
style_header_row(ws)
style_body(ws)
fit_widths(ws, [24, 90])

# ===== Sheet 9: 16 主题 & 5 活动主题 =====
ws = wb.create_sheet("9-主题清单")
ws.append(["序号", "媒体报道主题", "近似称谓 (摘自体系 docx)"])
THEME_DETAIL = {
    "美丽中国": "美丽中国建设、生态宜居",
    "综合治理": "生态保护、生态修复、生态环境综合治理、系统治理、环境治理",
    "绿色发展": "绿色低碳、低碳发展、绿色转型、零碳蓝碳",
    "双碳": "碳达峰碳中和、降污减碳、碳交易",
    "和谐共生": "地球生命共同体、绿色丝绸之路",
    "长江生态": "长江经济带生态保护、长江经济带、长江大保护、长江共抓大保护",
    "绿水青山": "绿水青山就是金山银山、两山",
    "制度建设": "生态文明制度、生态文明建设、生态文明体制改革",
    "资源节约": "资源节约集约利用、资源可循环",
    "污染防治攻坚战": "蓝天、碧水、净土保卫战",
    "清洁能源": "能源消费革命、新型能源体系、'无废城市'",
    "国家公园": "国家森林公园",
    "环保督察": "中央生态环境保护督察",
    "生物多样性": "生物多样性保护",
    "生态红线": "生态保护红线",
    "低碳经济": "绿色生活、低碳消费",
}
ORDER = ["美丽中国", "综合治理", "绿色发展", "双碳", "和谐共生", "长江生态", "绿水青山", "制度建设",
         "资源节约", "污染防治攻坚战", "清洁能源", "国家公园", "环保督察", "生物多样性", "生态红线", "低碳经济"]
for i, name in enumerate(ORDER, 1):
    ws.append([i, name, THEME_DETAIL[name]])
ws.append([])
ws.append(["序号", "公众活动主题", "说明"])
ACTIVITY = [
    "六五环境日活动", "8·15 全国生态日", "志愿服务活动",
    "环保设施向公众开放", "美丽重庆六进活动",
]
for i, name in enumerate(ACTIVITY, 1):
    ws.append([i, name, "—"])
style_header_row(ws)
ws.row_dimensions[len(ORDER) + 3].height = 18
header_row2 = len(ORDER) + 3
for c in ws[header_row2]:
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
style_body(ws)
fit_widths(ws, [6, 18, 60])

# 保存
Path(OUT_PATH).parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)
print(f"✓ 已写出: {OUT_PATH}")
print(f"  10 个 sheet:")
for name in wb.sheetnames:
    print(f"    - {name}")
