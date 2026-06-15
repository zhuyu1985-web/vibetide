#!/usr/bin/env python3
"""
基于 /tmp/ranking-v5-2025-full.json + /tmp/activities-2025.json
生成 docs/ranking-v5-2025-15sheet.xlsx —— 15 sheet (5 维度 × 3 子指标)

  1.1 中央-报道数量      | 1.2 中央-主题丰富度      | 1.3 中央-报道传播速度
  2.1 行业-报道数量      | 2.2 行业-主题丰富度      | 2.3 行业-报道传播速度
  3.1 市级-报道数量      | 3.2 市级-主题丰富度      | 3.3 市级-报道传播速度
  4.1 区县-报道数量      | 4.2 区县-主题丰富度      | 4.3 区县-报道传播速度
  5.1 公众-活动数量      | 5.2 公众-活动主题丰富度  | 5.3 公众-活动传播速度

每个 sheet 列设计:
  - 媒体类数量    : 排名, 区县, [16 主题各报道数], 报道总数, 区间化得分 (65-95)
  - 媒体类丰富度  : 排名, 区县, [16 主题各占比%], F 原始值, 区间化得分
  - 媒体类速度    : 排名, 区县, 报道总数, 发布天数, 速度原始, 区间化得分
  - 公众类数量    : 排名, 区县, [5 主题各场数], 活动总数, 区间化得分
  - 公众类丰富度  : 排名, 区县, [5 主题各占比%], F 原始值, 区间化得分
  - 公众类速度    : 排名, 区县, 活动总数, 首发日, 末发日, 跨度天, 速度原始, 区间化得分
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

JSON_FULL = "/tmp/ranking-v5-2025-full.json"
JSON_ACT = "/tmp/activities-2025.json"
OUT_PATH = "/Users/zhuyu/Developer/chinamcloud/vibetide/docs/ranking-v5-2025-15sheet.xlsx"

with open(JSON_FULL) as f:
    data = json.load(f)
with open(JSON_ACT) as f:
    activities = json.load(f)

ranked = data["ranked"]
raw_media = data["raw_media"]
scaled_media = data["scaled_media"]
raw_public = data["raw_public"]
scaled_public = data["scaled_public"]
topics = data["topics"]  # 16 个名字
activity_themes = data["activity_themes"]  # 5 个名字
N_TOPIC = len(topics)
N_ACT = len(activity_themes)

district_names = [r["name"] for r in ranked]
activity_by_district = {a["district"]: a for a in activities}

# 样式 ---------------------------------------------------------
wb = Workbook()
wb.remove(wb.active)
HEADER_FILL = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"),
)

def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

def style_body(ws, start_row=2):
    for ri in range(start_row, ws.max_row + 1):
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ri % 2 == 0:
                cell.fill = ALT_FILL

def fit_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# 各 tier 的中文标签
TIER_LABEL = {
    "central": "中央", "industry": "行业", "municipal": "市级",
    "district": "区县", "public": "公众",
}

# ============================================================
# 媒体类 3 个子 sheet 通用生成
# ============================================================
def gen_media_count_sheet(tier_key, sheet_name):
    """媒体类报道数量 sheet : 16 主题各报道数 + 总数 + 区间化得分"""
    ws = wb.create_sheet(sheet_name)
    headers = ["排名", "区县"] + topics + ["报道总数", "数量得分 (65-95)"]
    ws.append(headers)
    # 按报道总数降序
    rows = []
    for name in district_names:
        rm = raw_media[name][tier_key]
        sc = scaled_media[name][tier_key]
        rows.append((name, rm["topicCounts"], rm["count"], sc["count"]))
    rows.sort(key=lambda r: -r[2])
    for rank, (name, tc, total, score) in enumerate(rows, 1):
        row = [rank, name] + list(tc) + [total, round(score, 2)]
        ws.append(row)
    # 底部加平均行
    style_header_row(ws); style_body(ws)
    widths = [6, 14] + [10] * N_TOPIC + [12, 16]
    fit_widths(ws, widths)
    return ws

def gen_media_richness_sheet(tier_key, sheet_name):
    """媒体类丰富度 sheet : 16 主题占比% + F 原始 + 区间化得分"""
    ws = wb.create_sheet(sheet_name)
    headers = ["排名", "区县"] + [f"{t}占比%" for t in topics] + ["F 原始值", "丰富度得分 (65-95)"]
    ws.append(headers)
    rows = []
    for name in district_names:
        rm = raw_media[name][tier_key]
        sc = scaled_media[name][tier_key]
        total = max(rm["count"], 1)
        # topicCounts 是各 topic 在该 tier 的 annotation 命中数,百分比基于 sum(topicCounts)
        topic_total = max(sum(rm["topicCounts"]), 1)
        pcts = [round(c * 100 / topic_total, 2) for c in rm["topicCounts"]]
        rows.append((name, pcts, rm["richness"], sc["richness"]))
    rows.sort(key=lambda r: -r[2])
    for rank, (name, pcts, f_raw, score) in enumerate(rows, 1):
        row = [rank, name] + pcts + [round(f_raw, 4), round(score, 2)]
        ws.append(row)
    style_header_row(ws); style_body(ws)
    widths = [6, 14] + [11] * N_TOPIC + [12, 18]
    fit_widths(ws, widths)
    return ws

def gen_media_freq_sheet(tier_key, sheet_name):
    """媒体类速度 sheet : 报道总数 + 发布天数 + 速度原始 + 区间化得分"""
    ws = wb.create_sheet(sheet_name)
    ws.append(["排名", "区县", "报道总数", "发布天数", "速度原始 (报道数/天)", "速度得分 (65-95)"])
    rows = []
    for name in district_names:
        rm = raw_media[name][tier_key]
        sc = scaled_media[name][tier_key]
        rows.append((name, rm["count"], rm.get("days", 0), rm["freq"], sc["freq"]))
    rows.sort(key=lambda r: -r[3])
    for rank, (name, count, days, freq, score) in enumerate(rows, 1):
        ws.append([rank, name, count, days, round(freq, 4), round(score, 2)])
    style_header_row(ws); style_body(ws)
    fit_widths(ws, [6, 14, 12, 12, 20, 18])
    return ws

# ============================================================
# 公众类 3 个子 sheet
# ============================================================
def gen_public_count_sheet(sheet_name):
    ws = wb.create_sheet(sheet_name)
    ws.append(["排名", "区县"] + activity_themes + ["活动总数", "数量得分 (65-95)"])
    rows = []
    for name in district_names:
        rp = raw_public[name]
        sc = scaled_public[name]
        theme_counts = [(rp.get("themes", {}) or {}).get(t, 0) for t in activity_themes]
        rows.append((name, theme_counts, rp["count"], sc["count"]))
    rows.sort(key=lambda r: -r[2])
    for rank, (name, tc, total, score) in enumerate(rows, 1):
        ws.append([rank, name] + list(tc) + [total, round(score, 2)])
    style_header_row(ws); style_body(ws)
    fit_widths(ws, [6, 14] + [14] * N_ACT + [12, 18])
    return ws

def gen_public_richness_sheet(sheet_name):
    ws = wb.create_sheet(sheet_name)
    ws.append(["排名", "区县"] + [f"{t}占比%" for t in activity_themes] + ["F 原始值", "丰富度得分 (65-95)"])
    rows = []
    for name in district_names:
        rp = raw_public[name]
        sc = scaled_public[name]
        themes = rp.get("themes", {}) or {}
        total = max(sum(themes.values()), 1)
        pcts = [round(themes.get(t, 0) * 100 / total, 2) for t in activity_themes]
        rows.append((name, pcts, rp["richness"], sc["richness"]))
    rows.sort(key=lambda r: -r[2])
    for rank, (name, pcts, f_raw, score) in enumerate(rows, 1):
        ws.append([rank, name] + pcts + [round(f_raw, 4), round(score, 2)])
    style_header_row(ws); style_body(ws)
    fit_widths(ws, [6, 14] + [14] * N_ACT + [12, 18])
    return ws

def gen_public_freq_sheet(sheet_name):
    ws = wb.create_sheet(sheet_name)
    ws.append(["排名", "区县", "活动总数", "首发日", "末发日", "跨度天数", "速度原始 (场/天)", "速度得分 (65-95)"])
    rows = []
    for name in district_names:
        rp = raw_public[name]
        sc = scaled_public[name]
        rows.append((name, rp["count"], rp.get("firstDate", "-"), rp.get("lastDate", "-"),
                     rp.get("spanDays", 0), rp["freq"], sc["freq"]))
    rows.sort(key=lambda r: -r[5])
    for rank, (name, count, fd, ld, days, freq, score) in enumerate(rows, 1):
        ws.append([rank, name, count, fd, ld, days, round(freq, 4), round(score, 2)])
    style_header_row(ws); style_body(ws)
    fit_widths(ws, [6, 14, 12, 14, 14, 12, 20, 18])
    return ws

# ============================================================
# 生成 15 sheet
# ============================================================
SHEET_SPEC = [
    ("central", "1.1 中央-报道数量"),
    ("central", "1.2 中央-主题丰富度"),
    ("central", "1.3 中央-报道传播速度"),
    ("industry", "2.1 行业-报道数量"),
    ("industry", "2.2 行业-主题丰富度"),
    ("industry", "2.3 行业-报道传播速度"),
    ("municipal", "3.1 市级-报道数量"),
    ("municipal", "3.2 市级-主题丰富度"),
    ("municipal", "3.3 市级-报道传播速度"),
    ("district", "4.1 区县-报道数量"),
    ("district", "4.2 区县-主题丰富度"),
    ("district", "4.3 区县-报道传播速度"),
]
for tier_key, sheet_name in SHEET_SPEC:
    if sheet_name.endswith("数量"):
        gen_media_count_sheet(tier_key, sheet_name)
    elif sheet_name.endswith("丰富度"):
        gen_media_richness_sheet(tier_key, sheet_name)
    elif sheet_name.endswith("速度"):
        gen_media_freq_sheet(tier_key, sheet_name)

gen_public_count_sheet("5.1 公众-活动数量")
gen_public_richness_sheet("5.2 公众-活动主题丰富度")
gen_public_freq_sheet("5.3 公众-活动传播速度")

Path(OUT_PATH).parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)
print(f"✓ 已写出: {OUT_PATH}")
print(f"  {len(wb.sheetnames)} 个 sheet:")
for n in wb.sheetnames:
    print(f"    - {n}")
