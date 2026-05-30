#!/usr/bin/env python3
"""
基于新范围 /tmp/ranking-v5-2025-scope.json + media-scope-final.json 生成可验证 xlsx
新增 :
  - 00 总览 :加注明严格按 xlsx 收敛口径
  - 01 数据源清单 :列 94 家 媒体单位 (用户口径名 + 微博名 + ghid)
  - 02 数据范围审计 :加 dry-run 结果(保留率 99.1%, 政务=0, 开州=0)
  - 1.x-5.x 15 sheet
  - 99 综合汇总
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

JSON_SCOPE = "/tmp/ranking-v5-2025-scope.json"
JSON_SCOPE_DEF = "/tmp/media-scope-final.json"
JSON_DRYRUN = "/tmp/coverage-dryrun.json"
JSON_ACT = "/tmp/activities-2025.json"
OUT_PATH = "/Users/zhuyu/dev/chinamcloud/vibetide/docs/ranking-v5-2025-scope.xlsx"

with open(JSON_SCOPE) as f: data = json.load(f)
with open(JSON_SCOPE_DEF) as f: scope_def = json.load(f)
with open(JSON_DRYRUN) as f: dryrun = json.load(f)

ranked = data["ranked"]
raw_media = data["raw_media"]; scaled_media = data["scaled_media"]
raw_public = data["raw_public"]; scaled_public = data["scaled_public"]
topics = data["topics"]; activity_themes = data["activity_themes"]
N_TOPIC = len(topics); N_ACT = len(activity_themes)
TIER_WEIGHT = data["weights"]["tier"]; SUB_WEIGHT = data["weights"]["sub"]
SCALE_RANGE = data["weights"]["range"]
district_names = [r["name"] for r in ranked]

THEME_ALIAS = {
    "美丽中国": "美丽中国建设、生态宜居", "综合治理": "生态保护、生态修复、生态环境综合治理、系统治理、环境治理",
    "绿色发展": "绿色低碳、低碳发展、绿色转型、零碳蓝碳", "双碳": "碳达峰碳中和、降污减碳、碳交易",
    "和谐共生": "地球生命共同体、绿色丝绸之路", "长江生态": "长江经济带生态保护、长江经济带、长江大保护、长江共抓大保护",
    "绿水青山": "绿水青山就是金山银山、两山", "制度建设": "生态文明制度、生态文明建设、生态文明体制改革",
    "资源节约": "资源节约集约利用、资源可循环", "污染防治攻坚战": "蓝天、碧水、净土保卫战",
    "清洁能源": "能源消费革命、新型能源体系、'无废城市'", "国家公园": "国家森林公园",
    "环保督察": "中央生态环境保护督察", "生物多样性": "生物多样性保护",
    "生态红线": "生态保护红线", "低碳经济": "绿色生活、低碳消费",
}

wb = Workbook(); wb.remove(wb.active)
H_FILL = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
H_FONT = Font(bold=True, color="FFFFFF", size=11)
META_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
META_FONT = Font(bold=True, color="333333", size=10)
ALT_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
SECTION_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
SECTION_FONT = Font(bold=True, color="1B5E20", size=12)
WARN_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
BORDER = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))

def style_header_row(ws, row):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def style_data_rows(ws, start_row, end_row):
    for r in range(start_row, end_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if (r - start_row) % 2 == 1: cell.fill = ALT_FILL

def fit_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_meta_block(ws, meta_rows):
    for k, v in meta_rows:
        ws.append([k, v])
    for r in range(1, len(meta_rows) + 1):
        for c in range(1, 3):
            cell = ws.cell(row=r, column=c)
            cell.fill = META_FILL; cell.font = META_FONT if c == 1 else Font(size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = BORDER
    ws.append([])
    return len(meta_rows) + 2

# ============================================
# Sheet 00 :总览说明
# ============================================
ws = wb.create_sheet("00 总览说明")
overview = [
    ("【一、数据范围口径】", ""),
    ("数据范围", "严格按 /Users/zhuyu/Downloads/副本媒体站点名单-2(1).xlsx 的 94 家媒体单位收敛"),
    ("总媒体单位数", f"{len(scope_def['main_media_12']) + len(scope_def['district_rmt_41']) + len(scope_def['district_gov_41'])} 家 (主体 12 + 区县融媒 41 + 区县政务 41)"),
    ("数据窗口", "2025-01-01 ≤ published_at < 2026-01-01"),
    ("区县归并", "江北区 + 渝北区 → 两江新区(统一 39 区县口径)"),
    ("", ""),
    ("【二、本次范围内可用数据】(对比旧全字典范围)", ""),
    ("新范围 2025 items", f"{dryrun['summary']['items_in_scope']:,}"),
    ("旧范围 2025 items", f"{dryrun['summary']['items_old_full']:,}"),
    ("保留率", f"{dryrun['summary']['retention_pct']:.1f}%"),
    ("", ""),
    ("【三、需要说明的数据缺口】", ""),
    ("41 家区县政务号 items≈0", "字典中 government_self_media 的 public_account_names 全部为空,DB 中无对应稿件;本次政务部分计入 0(用户已知悉) — 区县媒体一级实际由 41 家融媒贡献"),
    ("", ""),
    ("【四、一级权重】(摘自指数体系 docx P34)", ""),
    ("中央媒体 (4 家)", "45%"),
    ("行业媒体 (2 家)", "25%"),
    ("市级媒体 (6 家)", "15%"),
    ("区县媒体 (82 家:41 融媒 + 41 政务)", "8%"),
    ("公众行为引导 (39 区县线下活动)", "7%"),
    ("", ""),
    ("【五、二级权重】", ""),
    ("报道/活动 数量", "40%"),
    ("报道/活动 主题丰富度", "30%"),
    ("报道/活动 传播速度", "30%"),
    ("", ""),
    ("【六、公式】", ""),
    ("主题丰富度 F", "F = 1/Σ|p_t − 1/N| ; N=16 (媒体) 或 5 (活动)"),
    ("传播速度", "freq = 报道/活动 总数 / 发布天数"),
    ("区间化", "min-max → [65, 95]"),
    ("一级得分", "tier = 数量×0.40 + 丰富度×0.30 + 速度×0.30"),
    ("综合得分", "综合 = 中央×0.45 + 行业×0.25 + 市级×0.15 + 区县×0.08 + 公众×0.07"),
]
for k, v in overview: ws.append([k, v])
fit_widths(ws, [44, 110])
for r in range(1, ws.max_row + 1):
    a = ws.cell(row=r, column=1); b = ws.cell(row=r, column=2)
    if isinstance(a.value, str) and a.value.startswith("【"):
        a.fill = SECTION_FILL; a.font = SECTION_FONT
        b.fill = SECTION_FILL; b.font = SECTION_FONT
    elif "缺口" in str(a.value or "") or "①" in str(a.value or "") or "②" in str(a.value or ""):
        a.fill = WARN_FILL; b.fill = WARN_FILL
    a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    a.border = BORDER; b.border = BORDER

# ============================================
# Sheet 01 :数据源清单 (94 家媒体)
# ============================================
ws = wb.create_sheet("01 数据源清单")
ws.append(["#", "类别", "媒体名(用户口径)", "区县归并", "公众号名", "公众号 ghid", "微博 UID", "网站域名"])
n = 0
# 12 主体
ws.append([f"=== 主体媒体 12 家(中央 4 + 市级 6 + 行业 2) ===", "", "", "", "", "", "", ""])
sec_row = ws.max_row
for c in range(1, 9):
    ws.cell(row=sec_row, column=c).fill = SECTION_FILL
    ws.cell(row=sec_row, column=c).font = SECTION_FONT
for u in scope_def["main_media_12"]:
    n += 1
    wc = "、".join(u.get("wechat_names") or []) or "—"
    web = "、".join(u.get("websites") or []) or "—"
    ws.append([n, u["tier"], u["name"], "—", wc, u.get("wechat_ghid") or "—", u.get("weibo_uid") or "—", web])
# 41 融媒
ws.append([f"=== 区县融媒体 41 家(江北/渝北→两江新区) ===", "", "", "", "", "", "", ""])
sec_row = ws.max_row
for c in range(1, 9):
    ws.cell(row=sec_row, column=c).fill = SECTION_FILL
    ws.cell(row=sec_row, column=c).font = SECTION_FONT
for u in scope_def["district_rmt_41"]:
    n += 1
    wc = "、".join(u.get("wechat_names") or []) or "—"
    web = "、".join(u.get("websites") or []) or "—"
    ws.append([n, "区县融媒", u["name"], u.get("district_normalized") or "—", wc, u.get("wechat_ghid") or "—", u.get("weibo_uid") or "—", web])
# 41 政务
ws.append([f"=== 区县政务新媒体 41 家(江北→两江,市生态环境局市级) ===", "", "", "", "", "", "", ""])
sec_row = ws.max_row
for c in range(1, 9):
    ws.cell(row=sec_row, column=c).fill = SECTION_FILL
    ws.cell(row=sec_row, column=c).font = SECTION_FONT
for u in scope_def["district_gov_41"]:
    n += 1
    wc = "、".join(u.get("wechat_names") or []) or "—"
    ws.append([n, "区县政务", u["name"], u.get("district_normalized") or "—", wc, u.get("wechat_ghid") or "—", u.get("weibo_uid") or "—", "—"])
fit_widths(ws, [5, 12, 28, 16, 32, 18, 14, 32])
for r in range(1, ws.max_row + 1):
    for c in range(1, 9):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 16 主题 + 5 活动
ws.append([])
ws.append([f"=== 16 个生态文明传播主题词 ===", "", "", "", "", "", "", ""])
sec_row = ws.max_row
for c in range(1, 9):
    ws.cell(row=sec_row, column=c).fill = SECTION_FILL
    ws.cell(row=sec_row, column=c).font = SECTION_FONT
for i, t in enumerate(topics, 1):
    ws.append([i, "主题词", t, "—", THEME_ALIAS.get(t, "—"), "—", "—", "—"])
ws.append([f"=== 5 个公众活动主题 ===", "", "", "", "", "", "", ""])
sec_row = ws.max_row
for c in range(1, 9):
    ws.cell(row=sec_row, column=c).fill = SECTION_FILL
    ws.cell(row=sec_row, column=c).font = SECTION_FONT
for i, t in enumerate(activity_themes, 1):
    ws.append([i, "活动主题", t, "—", "—", "—", "—", "—"])

# ============================================
# Sheet 02 :数据范围审计
# ============================================
ws = wb.create_sheet("02 数据范围审计")
audit = dryrun["summary"]
rows_02 = [
    ("【一、总体覆盖】", ""),
    ("旧范围 2025 items", f"{audit['items_old_full']:,}"),
    ("新范围 2025 items", f"{audit['items_in_scope']:,}"),
    ("保留率", f"{audit['retention_pct']:.1f}%"),
    ("匹配单位 / 总单位", f"{audit['matched_units']} / {audit['total_units']}"),
    ("匹配 outlet 数", f"{audit['matched_outlet_count']}"),
    ("", ""),
    ("【二、按 5 一级 tier 分布】", ""),
]
for k, v in rows_02: ws.append([k, v])

tier_labels = {"central": "中央 (45%)", "industry": "行业 (25%)", "municipal": "市级 (15%)", "district": "区县 (8%)"}
ws.append(["tier", "匹配单位", "匹配 outlet", "2025 items"])
hdr_row = ws.max_row
for tier_k in ["central", "industry", "municipal", "district"]:
    t = dryrun["by_tier"].get(tier_k, {})
    ws.append([tier_labels[tier_k], f"{t.get('matched_units',0)}/{t.get('total_units',0)}", t.get("matched_outlets", 0), f"{t.get('items',0):,}"])

ws.append([])
ws.append(["【三、按 39 区县分布(区县融媒 + 政务)】", ""])
ws.append(["区县", "融媒数", "融媒 items", "政务数", "政务 items", "区县合计"])
hdr_row2 = ws.max_row
STANDARD_39 = ["万州区","万盛经开区","两江新区","丰都县","九龙坡区","云阳县","北碚区","南岸区","南川区","合川区","垫江县","城口县","大渡口区","大足区","奉节县","巫山县","巫溪县","巴南区","开州区","彭水县","忠县","梁平区","武隆区","永川区","江津区","沙坪坝区","涪陵区","渝中区","潼南区","璧山区","石柱县","秀山县","科学城重庆高新区","綦江区","荣昌区","酉阳县","铜梁区","长寿区","黔江区"]
for d in STANDARD_39:
    a = dryrun["by_district"].get(d, {})
    total = a.get("rmt_items", 0) + a.get("gov_items", 0)
    ws.append([d, a.get("rmt_units", 0), a.get("rmt_items", 0), a.get("gov_units", 0), a.get("gov_items", 0), total])

fit_widths(ws, [28, 12, 14, 12, 14, 12])
for r in range(1, ws.max_row + 1):
    for c in range(1, 7):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if r in (hdr_row, hdr_row2):
            cell.fill = H_FILL; cell.font = H_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif isinstance(cell.value, str) and cell.value.startswith("【"):
            cell.fill = SECTION_FILL; cell.font = SECTION_FONT

# ============================================
# 1.x - 5.x : 15 个二级指标 sheet (沿用之前的设计)
# ============================================
TIER_LABEL = {"central": "中央媒体", "industry": "行业媒体", "municipal": "市级媒体", "district": "区县媒体"}
TIER_DB_FILTER = {
    "central": "已严格收敛: 央视/人民/光明/新华 4 家",
    "industry": "已严格收敛: 中环报 + 美丽重庆 2 家",
    "municipal": "已严格收敛: 上游/华龙/视界网/重庆日报/ichongqing/七一网 6 家",
    "district": "已严格收敛: 41 融媒 + 41 政务 = 82 家(政务实际无数据)",
}
TIER_W = {"central": "45%", "industry": "25%", "municipal": "15%", "district": "8%"}

def gen_media_count_sheet(tier_key, sheet_name):
    ws = wb.create_sheet(sheet_name)
    label = TIER_LABEL[tier_key]
    meta = [
        ("衡量指标", f"{label}-报道数量"),
        ("一级权重", TIER_W[tier_key]), ("二级权重", "40%"),
        ("数据源", TIER_DB_FILTER[tier_key]),
        ("过滤条件", "organization_id ; published_at ∈ [2025-01-01, 2026-01-01) ; 有 outlet_id"),
        ("原始值", "COUNT(DISTINCT collected_item_id) per (区县, tier) ; 按 16 主题分桶各显示命中数"),
        ("区间化", f"min-max → [{SCALE_RANGE[0]}, {SCALE_RANGE[1]}]"),
        ("排序", "按报道总数降序"),
    ]
    data_hdr = write_meta_block(ws, meta)
    ws.append(["排名", "区县"] + topics + ["报道总数 (原始)", f"数量得分 ({SCALE_RANGE[0]}-{SCALE_RANGE[1]})"])
    rows = []
    for name in district_names:
        rm = raw_media[name][tier_key]; sc = scaled_media[name][tier_key]
        rows.append((name, rm["topicCounts"], rm["count"], sc["count"]))
    rows.sort(key=lambda r: -r[2])
    for rank, (name, tc, total, score) in enumerate(rows, 1):
        ws.append([rank, name] + list(tc) + [total, round(score, 2)])
    style_header_row(ws, data_hdr); style_data_rows(ws, data_hdr + 1, ws.max_row)
    fit_widths(ws, [6, 14] + [10] * N_TOPIC + [14, 18])
    return ws

def gen_media_richness_sheet(tier_key, sheet_name):
    ws = wb.create_sheet(sheet_name)
    label = TIER_LABEL[tier_key]
    meta = [
        ("衡量指标", f"{label}-主题丰富度"), ("一级权重", TIER_W[tier_key]), ("二级权重", "30%"),
        ("数据源", TIER_DB_FILTER[tier_key]),
        ("公式", "F = 1 / Σ |p_t − 1/N| ; p_t = 某主题命中数 / 该(区县,tier)总命中 ; N=16"),
        ("中间列说明", "16 主题占比 (%) → |p_t − 1/16| 各项 → Σ → F = 1/Σ → 区间化得分"),
        ("区间化", f"min-max → [{SCALE_RANGE[0]}, {SCALE_RANGE[1]}]"),
        ("排序", "按 F 原始值降序"),
    ]
    data_hdr = write_meta_block(ws, meta)
    headers = ["排名", "区县"] + [f"{t}占比%" for t in topics] + [f"|{t}-1/16|" for t in topics] + ["Σ|p-1/16|", "F = 1/Σ", f"丰富度得分 ({SCALE_RANGE[0]}-{SCALE_RANGE[1]})"]
    ws.append(headers)
    rows = []; inv_n = 1 / N_TOPIC
    for name in district_names:
        rm = raw_media[name][tier_key]; sc = scaled_media[name][tier_key]
        tt = max(sum(rm["topicCounts"]), 1)
        pcts = [c / tt for c in rm["topicCounts"]]
        devs = [abs(p - inv_n) for p in pcts]
        sd = sum(devs); f_raw = (1 / sd) if sd > 0 else float("inf")
        rows.append((name, [round(p * 100, 2) for p in pcts], [round(d, 4) for d in devs], round(sd, 4),
                     round(f_raw, 4) if f_raw != float("inf") else "∞", round(sc["richness"], 2)))
    rows.sort(key=lambda r: (-(r[4] if isinstance(r[4], (int, float)) else 1e9)))
    for rank, (name, pcts, devs, sd, f_raw, score) in enumerate(rows, 1):
        ws.append([rank, name] + pcts + devs + [sd, f_raw, score])
    style_header_row(ws, data_hdr); style_data_rows(ws, data_hdr + 1, ws.max_row)
    fit_widths(ws, [6, 14] + [11] * N_TOPIC + [12] * N_TOPIC + [12, 12, 18])
    return ws

def gen_media_freq_sheet(tier_key, sheet_name):
    ws = wb.create_sheet(sheet_name)
    label = TIER_LABEL[tier_key]
    meta = [
        ("衡量指标", f"{label}-报道传播速度"), ("一级权重", TIER_W[tier_key]), ("二级权重", "30%"),
        ("数据源", TIER_DB_FILTER[tier_key]),
        ("公式", "freq = 报道总数 / 发布天数(COUNT DISTINCT YYYY-MM-DD)"),
        ("举例", "如某区某 tier 60 篇报道,分布在 30 个不同日期 → freq=60/30=2.0 篇/天"),
        ("区间化", f"min-max → [{SCALE_RANGE[0]}, {SCALE_RANGE[1]}]"),
        ("排序", "按速度原始值降序"),
    ]
    data_hdr = write_meta_block(ws, meta)
    ws.append(["排名", "区县", "报道总数", "发布天数", "速度原始 (报道/天)", f"速度得分 ({SCALE_RANGE[0]}-{SCALE_RANGE[1]})"])
    rows = []
    for name in district_names:
        rm = raw_media[name][tier_key]; sc = scaled_media[name][tier_key]
        rows.append((name, rm["count"], rm.get("days", 0), rm["freq"], sc["freq"]))
    rows.sort(key=lambda r: -r[3])
    for rank, (name, count, days, freq, score) in enumerate(rows, 1):
        ws.append([rank, name, count, days, round(freq, 4), round(score, 2)])
    style_header_row(ws, data_hdr); style_data_rows(ws, data_hdr + 1, ws.max_row)
    fit_widths(ws, [6, 14, 12, 18, 20, 18])
    return ws

def gen_public_count_sheet(sheet_name):
    ws = wb.create_sheet(sheet_name)
    meta = [
        ("衡量指标", "公众行为引导-活动数量"), ("一级权重", "7%"), ("二级权重", "40%"),
        ("数据源", "客户 Excel /Users/zhuyu/Downloads/副本2025年线下生态宣传活动统计表(1).xlsx 第 5-43 行"),
        ("原始值", "5 主题场数之和 = 活动总数"),
        ("区间化", f"min-max → [{SCALE_RANGE[0]}, {SCALE_RANGE[1]}]"),
        ("排序", "按活动总数降序"),
    ]
    data_hdr = write_meta_block(ws, meta)
    ws.append(["排名", "区县"] + activity_themes + ["活动总数 (原始)", f"数量得分 ({SCALE_RANGE[0]}-{SCALE_RANGE[1]})"])
    rows = []
    for name in district_names:
        rp = raw_public[name]; sc = scaled_public[name]
        tc = [(rp.get("themes", {}) or {}).get(t, 0) for t in activity_themes]
        rows.append((name, tc, rp["count"], sc["count"]))
    rows.sort(key=lambda r: -r[2])
    for rank, (name, tc, total, score) in enumerate(rows, 1):
        ws.append([rank, name] + list(tc) + [total, round(score, 2)])
    style_header_row(ws, data_hdr); style_data_rows(ws, data_hdr + 1, ws.max_row)
    fit_widths(ws, [6, 14] + [14] * N_ACT + [14, 18])
    return ws

def gen_public_richness_sheet(sheet_name):
    ws = wb.create_sheet(sheet_name)
    meta = [
        ("衡量指标", "公众行为引导-活动主题丰富度"), ("一级权重", "7%"), ("二级权重", "30%"),
        ("数据源", "客户 Excel 5 主题场数"),
        ("公式", "F = 1 / Σ |p_t − 1/5| ; p_t = 某主题场数 / 总场数 ; N=5"),
        ("中间列说明", "5 主题占比 (%) → |p_t − 1/5| → Σ → F = 1/Σ → 区间化"),
        ("区间化", f"min-max → [{SCALE_RANGE[0]}, {SCALE_RANGE[1]}]"),
        ("排序", "按 F 原始值降序"),
    ]
    data_hdr = write_meta_block(ws, meta)
    headers = ["排名", "区县"] + [f"{t}占比%" for t in activity_themes] + [f"|{t}-1/5|" for t in activity_themes] + ["Σ", "F = 1/Σ", f"丰富度得分"]
    ws.append(headers)
    rows = []; inv_n = 1 / N_ACT
    for name in district_names:
        rp = raw_public[name]; sc = scaled_public[name]
        themes = rp.get("themes", {}) or {}
        tt = max(sum(themes.values()), 1)
        pcts = [(themes.get(t, 0) / tt) for t in activity_themes]
        devs = [abs(p - inv_n) for p in pcts]
        sd = sum(devs); f_raw = (1 / sd) if sd > 0 else float("inf")
        rows.append((name, [round(p * 100, 2) for p in pcts], [round(d, 4) for d in devs], round(sd, 4),
                     round(f_raw, 4) if f_raw != float("inf") else "∞", round(sc["richness"], 2)))
    rows.sort(key=lambda r: (-(r[4] if isinstance(r[4], (int, float)) else 1e9)))
    for rank, (name, pcts, devs, sd, f_raw, score) in enumerate(rows, 1):
        ws.append([rank, name] + pcts + devs + [sd, f_raw, score])
    style_header_row(ws, data_hdr); style_data_rows(ws, data_hdr + 1, ws.max_row)
    fit_widths(ws, [6, 14] + [14] * N_ACT + [14] * N_ACT + [12, 12, 18])
    return ws

def gen_public_freq_sheet(sheet_name):
    ws = wb.create_sheet(sheet_name)
    meta = [
        ("衡量指标", "公众行为引导-活动传播速度"), ("一级权重", "7%"), ("二级权重", "30%"),
        ("数据源", "客户 Excel : 列 J(最晚日)/K(最早日)"),
        ("公式", "freq = 活动总数 / (最晚日 − 最早日 + 1)"),
        ("举例", "50 场,首场 2025-02-01,末场 2025-11-30,303 天 → freq=50/303≈0.165"),
        ("区间化", f"min-max → [{SCALE_RANGE[0]}, {SCALE_RANGE[1]}]"),
        ("排序", "按速度原始值降序"),
    ]
    data_hdr = write_meta_block(ws, meta)
    ws.append(["排名", "区县", "活动总数", "首发日", "末发日", "跨度天数", "速度原始 (场/天)", f"速度得分 ({SCALE_RANGE[0]}-{SCALE_RANGE[1]})"])
    rows = []
    for name in district_names:
        rp = raw_public[name]; sc = scaled_public[name]
        rows.append((name, rp["count"], rp.get("firstDate", "-"), rp.get("lastDate", "-"), rp.get("spanDays", 0), rp["freq"], sc["freq"]))
    rows.sort(key=lambda r: -r[5])
    for rank, (name, count, fd, ld, days, freq, score) in enumerate(rows, 1):
        ws.append([rank, name, count, fd, ld, days, round(freq, 4), round(score, 2)])
    style_header_row(ws, data_hdr); style_data_rows(ws, data_hdr + 1, ws.max_row)
    fit_widths(ws, [6, 14, 12, 14, 14, 14, 20, 18])
    return ws

SHEET_SPEC = [
    ("central", "1.1 中央-报道数量"), ("central", "1.2 中央-主题丰富度"), ("central", "1.3 中央-报道传播速度"),
    ("industry", "2.1 行业-报道数量"), ("industry", "2.2 行业-主题丰富度"), ("industry", "2.3 行业-报道传播速度"),
    ("municipal", "3.1 市级-报道数量"), ("municipal", "3.2 市级-主题丰富度"), ("municipal", "3.3 市级-报道传播速度"),
    ("district", "4.1 区县-报道数量"), ("district", "4.2 区县-主题丰富度"), ("district", "4.3 区县-报道传播速度"),
]
for tier_key, name in SHEET_SPEC:
    if name.endswith("数量"): gen_media_count_sheet(tier_key, name)
    elif name.endswith("丰富度"): gen_media_richness_sheet(tier_key, name)
    elif name.endswith("速度"): gen_media_freq_sheet(tier_key, name)
gen_public_count_sheet("5.1 公众-活动数量")
gen_public_richness_sheet("5.2 公众-活动主题丰富度")
gen_public_freq_sheet("5.3 公众-活动传播速度")

# Sheet 99
ws = wb.create_sheet("99 综合汇总")
meta = [
    ("【说明】", "5 维一级 × 权重 → 综合分,每个中间值可校验"),
    ("一级公式", "tier = 数量得分×0.40 + 丰富度得分×0.30 + 速度得分×0.30"),
    ("综合公式", "综合 = 中央×0.45 + 行业×0.25 + 市级×0.15 + 区县×0.08 + 公众×0.07"),
    ("数据范围", "严格按 xlsx 收敛: 中央 4 + 行业 2 + 市级 6 + 区县 82 + 公众 39 区县活动"),
]
data_hdr = write_meta_block(ws, meta)
ws.append(["排名", "区县", "中央 (45%)", "行业 (25%)", "市级 (15%)", "区县 (8%)", "公众 (7%)",
           "中央×0.45", "行业×0.25", "市级×0.15", "区县×0.08", "公众×0.07", "综合得分 (合计)"])
for r in ranked:
    name = r["name"]; c, i, m, d, p = r["central"], r["industry"], r["municipal"], r["district"], r["public"]
    pc, pi, pm, pd, pp = c*0.45, i*0.25, m*0.15, d*0.08, p*0.07
    composite = pc + pi + pm + pd + pp
    ws.append([r["rank"], name, round(c, 2), round(i, 2), round(m, 2), round(d, 2), round(p, 2),
               round(pc, 3), round(pi, 3), round(pm, 3), round(pd, 3), round(pp, 3), round(composite, 2)])
style_header_row(ws, data_hdr); style_data_rows(ws, data_hdr + 1, ws.max_row)
fit_widths(ws, [6, 16] + [12] * 5 + [12] * 5 + [16])

Path(OUT_PATH).parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)
print(f"✓ 已写出: {OUT_PATH}")
print(f"  共 {len(wb.sheetnames)} 个 sheet")
