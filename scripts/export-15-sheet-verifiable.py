#!/usr/bin/env python3
"""
基于 /tmp/ranking-v5-2025-full.json + /tmp/activities-2025.json + /tmp/audit-meta.json
生成 docs/ranking-v5-2025-verifiable.xlsx —— 19 sheet 可验证版

  00 总览说明        - 5 维 × 3 子的体系结构 + 权重 + 公式 + 数据源
  01 数据源清单      - 各 tier 真实 outlet 列表 + 16 主题词 + 5 活动主题
  02 数据范围审计    - DB 实际可用数据 + annotation 覆盖率
  1.1~5.3 (15 sheet)  - 每 sheet 顶部 8 行算法说明区 + 推导明细
  99 综合汇总        - 5 维加权 → 综合分,可逐步校验

设计原则 :
  - 每个数据点都能追溯到 (DB 表, 字段, 过滤条件, 公式)
  - 丰富度 sheet 展开 |p_t - 1/N| 推导
  - 速度 sheet 展开 总数 / 发布天数 = 速度
  - 区间化方法在 00 sheet 和每个 sheet 顶部都重复说明
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

JSON_FULL = "/tmp/ranking-v5-2025-full.json"
JSON_ACT = "/tmp/activities-2025.json"
JSON_AUDIT = "/tmp/audit-meta.json"
OUT_PATH = "/Users/zhuyu/dev/chinamcloud/vibetide/docs/ranking-v5-2025-verifiable.xlsx"

with open(JSON_FULL) as f:
    data = json.load(f)
with open(JSON_ACT) as f:
    activities = json.load(f)
with open(JSON_AUDIT) as f:
    audit = json.load(f)

ranked = data["ranked"]
raw_media = data["raw_media"]
scaled_media = data["scaled_media"]
raw_public = data["raw_public"]
scaled_public = data["scaled_public"]
topics = data["topics"]
activity_themes = data["activity_themes"]
N_TOPIC = len(topics)
N_ACT = len(activity_themes)
TIER_WEIGHT = data["weights"]["tier"]
SUB_WEIGHT = data["weights"]["sub"]
SCALE_RANGE = data["weights"]["range"]

district_names = [r["name"] for r in ranked]
finals_by_name = {r["name"]: r for r in ranked}

# 主题近似称谓 (摘自体系 docx 表 3-1)
THEME_ALIAS = {
    "美丽中国": "美丽中国建设、生态宜居",
    "综合治理": "生态保护、生态修复、生态环境综合治理、系统治理、环境治理",
    "绿色发展": "绿色低碳、低碳发展、绿色转型、零碳蓝碳",
    "双碳": "碳达峰碳中和、降污减碳、碳交易",
    "和谐共生": "地球生命共同体、绿色丝绸之路",
    "长江生态": "长江经济带生态保护、长江经济带、长江大保护、长江共抓大保护",
    "绿水青山": "绿水青山就是金山银山、两山",
    "制度建设": "生态文明制度、生态文明建设、生态文明体制改革",
    "资源节约": "资源节约集约利用、资源可循环",
    "污染防治攻坚战": "蓝天、碧水、净土保卫战",
    "清洁能源": "能源消费革命、新型能源体系、'无废城市'",
    "国家公园": "国家森林公园",
    "环保督察": "中央生态环境保护督察",
    "生物多样性": "生物多样性保护",
    "生态红线": "生态保护红线",
    "低碳经济": "绿色生活、低碳消费",
}

# === 样式 ===
wb = Workbook()
wb.remove(wb.active)
H_FILL = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
H_FONT = Font(bold=True, color="FFFFFF", size=11)
META_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
META_FONT = Font(bold=True, color="333333", size=10)
ALT_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
SECTION_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
SECTION_FONT = Font(bold=True, color="1B5E20", size=12)
BORDER = Border(
    left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"),
)

def style_meta_block(ws, start_row, end_row):
    for r in range(start_row, end_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = META_FILL; cell.font = META_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = BORDER

def style_header_row(ws, row):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def style_data_rows(ws, start_row, end_row):
    for r in range(start_row, end_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if (r - start_row) % 2 == 1:
                cell.fill = ALT_FILL

def fit_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# =============================================================
# Sheet 00 : 总览说明
# =============================================================
ws = wb.create_sheet("00 总览说明")
rows = [
    ("【一、体系结构】", ""),
    ("一级指标", "5 个 :中央媒体 / 行业媒体 / 市级媒体 / 区县媒体 / 公众行为引导"),
    ("二级指标", "3 个 :报道数量(活动数量) / 报道主题丰富度(活动主题丰富度) / 报道传播速度(活动传播速度)"),
    ("总指标数", "5 × 3 = 15 个二级指标"),
    ("数据窗口", "2025-01-01 ≤ published_at < 2026-01-01"),
    ("", ""),
    ("【二、一级权重】(摘自体系 docx P34)", ""),
    ("中央媒体传播指数", "45%"),
    ("行业媒体传播指数", "25%"),
    ("市级媒体传播指数", "15%"),
    ("区县媒体传播指数", "8%"),
    ("公众行为引导指数", "7%"),
    ("权重合计", "100%"),
    ("", ""),
    ("【三、二级权重】(摘自体系 docx P34)", ""),
    ("报道数量 / 活动数量", "40%"),
    ("报道主题丰富度 / 活动主题丰富度", "30%"),
    ("报道传播速度 / 活动传播速度", "30%"),
    ("权重合计", "100%"),
    ("", ""),
    ("【四、二级指标计算公式】(摘自体系 docx P55-61)", ""),
    ("(1) 报道/活动 数量", "对原始数据去重后,按 39 区县分别统计该 (区县, tier) 下命中稿件总数 / 活动总数"),
    ("(2) 主题丰富度 F", "F = 1 / Σ |p_t − 1/N| ; p_t 为某主题占该 (区县, tier) 报道总数的比例 ; N=16(媒体) 或 5(活动) ; F>0 越大越好,表示该 (区县, tier) 报道能均匀覆盖各主题"),
    ("(3) 传播速度", "freq = 报道/活动 总数 / 有发布的不同日数(按 YYYY-MM-DD 去重计) ; 活动用 (最晚日 − 最早日 + 1) 天"),
    ("", ""),
    ("【五、区间化方法】(摘自体系 docx P40)", ""),
    ("方法", "将量纲不同的 15 个二级指标的原始数据,按区县间 min-max 标准化到 [65, 95] 区间"),
    ("公式", "score = 65 + (raw − min_district) / (max_district − min_district) × 30"),
    ("说明", "每个二级指标在 39 区县间独立做一次区间化 ;一级指标不再二次区间化,直接由二级加权得"),
    ("", ""),
    ("【六、一级指标合成】", ""),
    ("一级 = 数量得分×0.40 + 丰富度得分×0.30 + 速度得分×0.30", ""),
    ("", ""),
    ("【七、综合得分】", ""),
    ("综合 = 中央×0.45 + 行业×0.25 + 市级×0.15 + 区县×0.08 + 公众×0.07", ""),
    ("", ""),
    ("【八、本表所有数据的来源】", ""),
    ("媒体类原始数据 (中央/行业/市级/区县)", "DB : collected_items × research_collected_item_districts × research_collected_item_topics × media_outlet_dictionary  (按 outlet_tier 分桶)"),
    ("公众类原始数据", "Excel : /Users/zhuyu/Downloads/副本2025年线下生态宣传活动统计表(1).xlsx  39 行 × 5 主题场数 + 时间窗"),
    ("主题词字典", "DB : research_topics (16 行)"),
    ("39 区县字典", "DB : research_cq_districts (39 行)"),
    ("annotation 命中规则", "src/lib/research/topic-matcher.ts 和 district-matcher.ts ;基于稿件标题+正文+OCR+ASR+tags+主题词关键词匹配"),
    ("", ""),
    ("【九、可验证性】", ""),
    ("Sheet 01", "列出本次计算用到的全部 outlet 名(中央 12 / 行业 32 / 市级 32 / 区县 84),可去 DB 查"),
    ("Sheet 02", "列出 DB 范围内的实际可用数据,可逐项 SQL 复现"),
    ("Sheet 1.x-5.x", "每个 sheet 顶部 8 行算法说明区,公式 + 数据源 + 推导步骤"),
    ("Sheet 99", "5 维一级 × 权重 → 综合分,可手工 ×+÷ 校验"),
]
for k, v in rows:
    ws.append([k, v])
fit_widths(ws, [50, 110])
for r in range(1, ws.max_row + 1):
    cell_a = ws.cell(row=r, column=1)
    cell_b = ws.cell(row=r, column=2)
    text = (cell_a.value or "")
    if isinstance(text, str) and text.startswith("【"):
        cell_a.fill = SECTION_FILL; cell_a.font = SECTION_FONT
        cell_b.fill = SECTION_FILL; cell_b.font = SECTION_FONT
    cell_a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell_b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell_a.border = BORDER; cell_b.border = BORDER

# =============================================================
# Sheet 01 : 数据源清单 - 各 tier outlet + 主题词
# =============================================================
ws = wb.create_sheet("01 数据源清单")
TIER_LABEL_MAP = {
    "central": "中央 (权重 45%) - DB outlet_tier='central'",
    "industry": "行业 (权重 25%) - DB outlet_tier='industry'",
    "provincial_municipal": "市级 (权重 15%) - DB outlet_tier='provincial_municipal'",
    "district_media": "区县(融媒体) (权重 8% 一半) - DB outlet_tier='district_media'",
    "government_self_media": "区县(生态环境政务) (权重 8% 一半) - DB outlet_tier='government_self_media'",
}
ws.append(["#", "类别", "outlet 名称 (DB 字典)", "区域", "公众号别名 (前 3 个)"])
n = 0
for tier, label in TIER_LABEL_MAP.items():
    rows_tier = audit["outlets"].get(tier, [])
    # 类别分隔行
    n += 1
    ws.append([f"=== {label} - 共 {len(rows_tier)} 家 ===", "", "", "", ""])
    sec_row = ws.max_row
    for c in range(1, 6):
        ws.cell(row=sec_row, column=c).fill = SECTION_FILL
        ws.cell(row=sec_row, column=c).font = SECTION_FONT
    for o in rows_tier:
        n += 1
        pa = ", ".join((o.get("pa") or [])[:3])
        if len(o.get("pa") or []) > 3:
            pa += " ..."
        ws.append([n, label.split(" ")[0], o["name"], o.get("region") or "-", pa])
# 16 主题
ws.append([])
ws.append(["=== 16 个生态文明传播主题词 (DB : research_topics) ==="])
sec_row = ws.max_row
for c in range(1, 6):
    ws.cell(row=sec_row, column=c).fill = SECTION_FILL
    ws.cell(row=sec_row, column=c).font = SECTION_FONT
ws.append(["#", "主题词", "近似称谓 (体系 docx 表 3-1)", "", ""])
hdr_row = ws.max_row
for i, t in enumerate(topics, 1):
    ws.append([i, t, THEME_ALIAS.get(t, "—"), "", ""])
# 5 活动主题
ws.append([])
ws.append(["=== 5 个公众活动主题 (体系 docx P59 / 客户 Excel 列头) ==="])
sec_row = ws.max_row
for c in range(1, 6):
    ws.cell(row=sec_row, column=c).fill = SECTION_FILL
    ws.cell(row=sec_row, column=c).font = SECTION_FONT
for i, t in enumerate(activity_themes, 1):
    ws.append([i, t, "—", "", ""])
fit_widths(ws, [6, 36, 36, 12, 36])
for r in range(1, ws.max_row + 1):
    for c in range(1, 6):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# =============================================================
# Sheet 02 : 数据范围审计
# =============================================================
ws = wb.create_sheet("02 数据范围审计")
ws.append(["【一、collected_items 表数据状况】"])
ws.append(["指标", "数值", "SQL 复现"])
items = audit["audit"]["items"]
ws.append(["org_id", audit.get("org_id", "—"), "SELECT id FROM organizations LIMIT 1"])
ws.append(["稿件总条数 (org-scoped)", items["total"], "SELECT COUNT(*) FROM collected_items WHERE organization_id = ?"])
ws.append(["有 outlet_id 的条数", items["with_outlet"], "SELECT COUNT(*) FROM collected_items WHERE organization_id = ? AND outlet_id IS NOT NULL"])
ws.append(["2025 年内的条数", items["in_2025"], "SELECT COUNT(*) FROM collected_items WHERE organization_id = ? AND published_at >= '2025-01-01' AND published_at < '2026-01-01'"])
ws.append(["2025 + 有 outlet (本次计算输入)", items["in_2025_with_outlet"], "上面两个条件 AND 取交集"])
ws.append([""])
ws.append(["【二、按 tier 分组 (2025 + 有 outlet)】"])
ws.append(["tier", "稿件数", ""])
for r in audit["audit"]["by_tier"]:
    ws.append([r["outlet_tier"], r["n"], ""])
ws.append([""])
ws.append(["【三、annotation 命中情况】"])
ws.append(["指标", "数值", "SQL 复现"])
ann = audit["audit"]["annotations"]
ws.append(["有 topic annotation 的稿件数", ann["items_w_topic"], "SELECT COUNT(DISTINCT collected_item_id) FROM research_collected_item_topics"])
ws.append(["有 district annotation 的稿件数", ann["items_w_district"], "SELECT COUNT(DISTINCT collected_item_id) FROM research_collected_item_districts"])
ws.append(["topic annotation 总条数", ann["topic_n"], "SELECT COUNT(*) FROM research_collected_item_topics"])
ws.append(["district annotation 总条数", ann["district_n"], "SELECT COUNT(*) FROM research_collected_item_districts"])
ws.append([""])
ws.append(["【四、39 区县的稿件数分布 (有 district annotation 的)】"])
ws.append(["区县", "稿件数", "占比"])
total_d = sum(r["n"] for r in audit["audit"]["by_district"])
for r in audit["audit"]["by_district"]:
    pct = round(r["n"] * 100 / max(total_d, 1), 2)
    ws.append([r["name"], r["n"], f"{pct}%"])
fit_widths(ws, [42, 28, 100])
for r in range(1, ws.max_row + 1):
    for c in range(1, 4):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if r == 1 or (isinstance(cell.value, str) and (cell.value or "").startswith("【")):
            cell.fill = SECTION_FILL; cell.font = SECTION_FONT

# =============================================================
# 通用 : sheet 顶部 META 说明区 (8 行) - 写完返回数据起始行
# =============================================================
def write_meta_block(ws, meta_rows):
    """写算法说明区,返回 data_header_row (表头所在行号)"""
    for k, v in meta_rows:
        ws.append([k, v])
    # 给元数据行加样式
    for r in range(1, len(meta_rows) + 1):
        for c in range(1, 3):
            cell = ws.cell(row=r, column=c)
            cell.fill = META_FILL
            cell.font = META_FONT if (cell.value or "").startswith(("衡量","数据","公式","区间化","二级权重","一级权重")) else Font(size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = BORDER
    # 空一行
    ws.append([])
    return len(meta_rows) + 2  # 下一行就是表头

# =============================================================
# 1.x ~ 4.x : 媒体类 12 个 sheet
# =============================================================
TIER_LABEL = {
    "central": "中央媒体",
    "industry": "行业媒体",
    "municipal": "市级媒体",
    "district": "区县媒体",
}
TIER_DB_FILTER = {
    "central": "outlet_tier = 'central'",
    "industry": "outlet_tier = 'industry'",
    "municipal": "outlet_tier = 'provincial_municipal'",
    "district": "outlet_tier IN ('district_media', 'government_self_media')",
}
TIER_W = {"central": "45%", "industry": "25%", "municipal": "15%", "district": "8%"}

def gen_media_count_sheet(tier_key, sheet_name):
    ws = wb.create_sheet(sheet_name)
    label = TIER_LABEL[tier_key]
    meta = [
        (f"衡量指标", f"{label}-报道数量"),
        (f"一级权重", f"{TIER_W[tier_key]} (本一级在综合分中的占比)"),
        (f"二级权重", "40% (本子指标在一级中的占比)"),
        (f"数据源", f"DB collected_items × research_collected_item_topics × media_outlet_dictionary  ({TIER_DB_FILTER[tier_key]})"),
        (f"过滤条件", "organization_id 一致 ;  published_at ∈ [2025-01-01, 2026-01-01) ; 有 outlet_id"),
        (f"原始值计算", "对每个 (区县, tier) 组,COUNT(DISTINCT collected_item_id) 为报道总数;按 16 主题分桶各显示命中数"),
        (f"区间化方法", f"在 39 区县间 min-max 标准化到 [{SCALE_RANGE[0]}, {SCALE_RANGE[1]}],公式 65 + (raw-min)/(max-min) × 30"),
        (f"排序", "按本指标的报道总数降序"),
    ]
    data_hdr = write_meta_block(ws, meta)
    headers = ["排名", "区县"] + topics + ["报道总数 (原始)", f"数量得分 ({SCALE_RANGE[0]}-{SCALE_RANGE[1]})"]
    ws.append(headers)
    for name in sorted(district_names, key=lambda d: -raw_media[d][tier_key]["count"]):
        rm = raw_media[name][tier_key]
        sc = scaled_media[name][tier_key]
        pass  # 留位
    # 实际写入
    rows = []
    for name in district_names:
        rm = raw_media[name][tier_key]
        sc = scaled_media[name][tier_key]
        rows.append((name, rm["topicCounts"], rm["count"], sc["count"]))
    rows.sort(key=lambda r: -r[2])
    for rank, (name, tc, total, score) in enumerate(rows, 1):
        ws.append([rank, name] + list(tc) + [total, round(score, 2)])
    style_header_row(ws, data_hdr)
    style_data_rows(ws, data_hdr + 1, ws.max_row)
    widths = [6, 14] + [10] * N_TOPIC + [14, 18]
    fit_widths(ws, widths)
    return ws

def gen_media_richness_sheet(tier_key, sheet_name):
    ws = wb.create_sheet(sheet_name)
    label = TIER_LABEL[tier_key]
    meta = [
        ("衡量指标", f"{label}-主题丰富度"),
        ("一级权重", f"{TIER_W[tier_key]}"),
        ("二级权重", "30%"),
        ("数据源", f"DB ({TIER_DB_FILTER[tier_key]}) 的报道,按 16 主题分桶后算各主题占比"),
        ("公式", "F = 1 / Σ |p_t − 1/N| ;  p_t = 某主题命中数 / 该 (区县, tier) 16 主题命中总数 ;  N=16"),
        ("中间列说明", "16 主题占比 (%) → |p_t − 1/16| 各项 → Σ → F = 1/Σ → 区间化得分"),
        ("区间化方法", f"min-max → [{SCALE_RANGE[0]}, {SCALE_RANGE[1]}]"),
        ("排序", "按 F 原始值降序"),
    ]
    data_hdr = write_meta_block(ws, meta)
    # 列 : 排名/区县/16 主题占比%/16 |p-1/16|/Σ|p-1/16|/F 原始/得分
    headers = ["排名", "区县"] \
        + [f"{t}占比%" for t in topics] \
        + [f"|{t.replace('占比','')}-1/16|" for t in topics] \
        + ["Σ|p-1/16|", "F = 1/Σ", f"丰富度得分 ({SCALE_RANGE[0]}-{SCALE_RANGE[1]})"]
    ws.append(headers)
    rows = []
    inv_n = 1 / N_TOPIC
    for name in district_names:
        rm = raw_media[name][tier_key]
        sc = scaled_media[name][tier_key]
        topic_total = max(sum(rm["topicCounts"]), 1)
        pcts = [c / topic_total for c in rm["topicCounts"]]
        devs = [abs(p - inv_n) for p in pcts]
        sum_dev = sum(devs)
        f_raw = (1 / sum_dev) if sum_dev > 0 else float("inf")
        rows.append((
            name,
            [round(p * 100, 2) for p in pcts],
            [round(d, 4) for d in devs],
            round(sum_dev, 4),
            round(f_raw, 4) if f_raw != float("inf") else "∞",
            round(sc["richness"], 2),
        ))
    rows.sort(key=lambda r: (-(r[4] if isinstance(r[4], (int, float)) else 1e9)))
    for rank, (name, pcts, devs, sum_dev, f_raw, score) in enumerate(rows, 1):
        ws.append([rank, name] + pcts + devs + [sum_dev, f_raw, score])
    style_header_row(ws, data_hdr)
    style_data_rows(ws, data_hdr + 1, ws.max_row)
    widths = [6, 14] + [11] * N_TOPIC + [12] * N_TOPIC + [12, 12, 18]
    fit_widths(ws, widths)
    return ws

def gen_media_freq_sheet(tier_key, sheet_name):
    ws = wb.create_sheet(sheet_name)
    label = TIER_LABEL[tier_key]
    meta = [
        ("衡量指标", f"{label}-报道传播速度"),
        ("一级权重", f"{TIER_W[tier_key]}"),
        ("二级权重", "30%"),
        ("数据源", f"DB ({TIER_DB_FILTER[tier_key]}) 的报道"),
        ("公式", "freq = 报道总数 / 发布天数  (发布天数 = COUNT(DISTINCT to_char(published_at, 'YYYY-MM-DD')))"),
        ("举例", "如某区某 tier 一年内有 60 篇报道,分布在 30 个不同日期 → freq = 60/30 = 2.0 (篇/天)"),
        ("区间化方法", f"min-max → [{SCALE_RANGE[0]}, {SCALE_RANGE[1]}]"),
        ("排序", "按速度原始值降序"),
    ]
    data_hdr = write_meta_block(ws, meta)
    headers = ["排名", "区县", "报道总数", "发布天数 (不同日期数)", "速度原始 (报道/天)", f"速度得分 ({SCALE_RANGE[0]}-{SCALE_RANGE[1]})"]
    ws.append(headers)
    rows = []
    for name in district_names:
        rm = raw_media[name][tier_key]
        sc = scaled_media[name][tier_key]
        rows.append((name, rm["count"], rm.get("days", 0), rm["freq"], sc["freq"]))
    rows.sort(key=lambda r: -r[3])
    for rank, (name, count, days, freq, score) in enumerate(rows, 1):
        ws.append([rank, name, count, days, round(freq, 4), round(score, 2)])
    style_header_row(ws, data_hdr)
    style_data_rows(ws, data_hdr + 1, ws.max_row)
    fit_widths(ws, [6, 14, 12, 18, 20, 18])
    return ws

# =============================================================
# 5.x : 公众类 3 个 sheet
# =============================================================
def gen_public_count_sheet(sheet_name):
    ws = wb.create_sheet(sheet_name)
    meta = [
        ("衡量指标", "公众行为引导-活动数量"),
        ("一级权重", "7%"),
        ("二级权重", "40%"),
        ("数据源", "客户 Excel : /Users/zhuyu/Downloads/副本2025年线下生态宣传活动统计表(1).xlsx 第 5-43 行"),
        ("过滤条件", "无 (Excel 已是 39 区县全集)"),
        ("原始值计算", "对每行 5 个主题场数求和 = 该区县活动总数"),
        ("区间化方法", f"min-max → [{SCALE_RANGE[0]}, {SCALE_RANGE[1]}]"),
        ("排序", "按活动总数降序"),
    ]
    data_hdr = write_meta_block(ws, meta)
    headers = ["排名", "区县"] + activity_themes + ["活动总数 (原始)", f"数量得分 ({SCALE_RANGE[0]}-{SCALE_RANGE[1]})"]
    ws.append(headers)
    rows = []
    for name in district_names:
        rp = raw_public[name]
        sc = scaled_public[name]
        theme_counts = [(rp.get("themes", {}) or {}).get(t, 0) for t in activity_themes]
        rows.append((name, theme_counts, rp["count"], sc["count"]))
    rows.sort(key=lambda r: -r[2])
    for rank, (name, tc, total, score) in enumerate(rows, 1):
        ws.append([rank, name] + list(tc) + [total, round(score, 2)])
    style_header_row(ws, data_hdr)
    style_data_rows(ws, data_hdr + 1, ws.max_row)
    fit_widths(ws, [6, 14] + [14] * N_ACT + [14, 18])
    return ws

def gen_public_richness_sheet(sheet_name):
    ws = wb.create_sheet(sheet_name)
    meta = [
        ("衡量指标", "公众行为引导-活动主题丰富度"),
        ("一级权重", "7%"),
        ("二级权重", "30%"),
        ("数据源", "客户 Excel 5 主题场数 (同 5.1 sheet)"),
        ("公式", "F = 1 / Σ |p_t − 1/5| ;  p_t = 某主题场数/总场数 ;  N=5"),
        ("中间列说明", "5 主题占比 (%) → |p_t − 1/5| 各项 → Σ → F = 1/Σ → 区间化得分"),
        ("区间化方法", f"min-max → [{SCALE_RANGE[0]}, {SCALE_RANGE[1]}]"),
        ("排序", "按 F 原始值降序"),
    ]
    data_hdr = write_meta_block(ws, meta)
    headers = ["排名", "区县"] \
        + [f"{t}占比%" for t in activity_themes] \
        + [f"|{t}-1/5|" for t in activity_themes] \
        + ["Σ|p-1/5|", "F = 1/Σ", f"丰富度得分 ({SCALE_RANGE[0]}-{SCALE_RANGE[1]})"]
    ws.append(headers)
    rows = []
    inv_n = 1 / N_ACT
    for name in district_names:
        rp = raw_public[name]
        sc = scaled_public[name]
        themes = rp.get("themes", {}) or {}
        total = max(sum(themes.values()), 1)
        pcts = [(themes.get(t, 0) / total) for t in activity_themes]
        devs = [abs(p - inv_n) for p in pcts]
        sum_dev = sum(devs)
        f_raw = (1 / sum_dev) if sum_dev > 0 else float("inf")
        rows.append((
            name,
            [round(p * 100, 2) for p in pcts],
            [round(d, 4) for d in devs],
            round(sum_dev, 4),
            round(f_raw, 4) if f_raw != float("inf") else "∞",
            round(sc["richness"], 2),
        ))
    rows.sort(key=lambda r: (-(r[4] if isinstance(r[4], (int, float)) else 1e9)))
    for rank, (name, pcts, devs, sum_dev, f_raw, score) in enumerate(rows, 1):
        ws.append([rank, name] + pcts + devs + [sum_dev, f_raw, score])
    style_header_row(ws, data_hdr)
    style_data_rows(ws, data_hdr + 1, ws.max_row)
    fit_widths(ws, [6, 14] + [14] * N_ACT + [14] * N_ACT + [12, 12, 18])
    return ws

def gen_public_freq_sheet(sheet_name):
    ws = wb.create_sheet(sheet_name)
    meta = [
        ("衡量指标", "公众行为引导-活动传播速度"),
        ("一级权重", "7%"),
        ("二级权重", "30%"),
        ("数据源", "客户 Excel : 列 J(最晚日期) 和列 K(最早日期),用 excel 日期序号 → 实际日期"),
        ("公式", "freq = 活动总数 / (最晚日 − 最早日 + 1)  (单位 : 场/天)"),
        ("举例", "如某区一年 50 场活动,首场 2025-02-01,末场 2025-11-30,共 303 天 → freq = 50/303 ≈ 0.165"),
        ("区间化方法", f"min-max → [{SCALE_RANGE[0]}, {SCALE_RANGE[1]}]"),
        ("排序", "按速度原始值降序"),
    ]
    data_hdr = write_meta_block(ws, meta)
    headers = ["排名", "区县", "活动总数", "首发日", "末发日", "跨度天数 (含)", "速度原始 (场/天)", f"速度得分 ({SCALE_RANGE[0]}-{SCALE_RANGE[1]})"]
    ws.append(headers)
    rows = []
    for name in district_names:
        rp = raw_public[name]
        sc = scaled_public[name]
        rows.append((name, rp["count"], rp.get("firstDate", "-"), rp.get("lastDate", "-"), rp.get("spanDays", 0), rp["freq"], sc["freq"]))
    rows.sort(key=lambda r: -r[5])
    for rank, (name, count, fd, ld, days, freq, score) in enumerate(rows, 1):
        ws.append([rank, name, count, fd, ld, days, round(freq, 4), round(score, 2)])
    style_header_row(ws, data_hdr)
    style_data_rows(ws, data_hdr + 1, ws.max_row)
    fit_widths(ws, [6, 14, 12, 14, 14, 14, 20, 18])
    return ws

# =============================================================
# 生成 15 个二级指标 sheet
# =============================================================
SHEET_SPEC = [
    ("central", "1.1 中央-报道数量"),
    ("central", "1.2 中央-主题丰富度"),
    ("central", "1.3 中央-报道传播速度"),
    ("industry", "2.1 行业-报道数量"),
    ("industry", "2.2 行业-主题丰富度"),
    ("industry", "2.3 行业-报道传播速度"),
    ("municipal", "3.1 市级-报道数量"),
    ("municipal", "3.2 市级-主题丰富度"),
    ("municipal", "3.3 市级-报道传播速度"),
    ("district", "4.1 区县-报道数量"),
    ("district", "4.2 区县-主题丰富度"),
    ("district", "4.3 区县-报道传播速度"),
]
for tier_key, name in SHEET_SPEC:
    if name.endswith("数量"):
        gen_media_count_sheet(tier_key, name)
    elif name.endswith("丰富度"):
        gen_media_richness_sheet(tier_key, name)
    elif name.endswith("速度"):
        gen_media_freq_sheet(tier_key, name)

gen_public_count_sheet("5.1 公众-活动数量")
gen_public_richness_sheet("5.2 公众-活动主题丰富度")
gen_public_freq_sheet("5.3 公众-活动传播速度")

# =============================================================
# Sheet 99 : 综合汇总 (五维加权 → 综合)
# =============================================================
ws = wb.create_sheet("99 综合汇总")
meta = [
    ("【说明】", "本 sheet 把 5 维一级分按权重加权得到综合分,所有中间值可手工 ×+÷ 校验"),
    ("一级公式", "tier = 数量得分×0.40 + 丰富度得分×0.30 + 速度得分×0.30"),
    ("综合公式", "综合 = 中央×0.45 + 行业×0.25 + 市级×0.15 + 区县×0.08 + 公众×0.07"),
    ("数据来源", "本 sheet 行中的 '中央/行业/市级/区县/公众' 5 列等于对应 Sheet 1.x-5.x 末列的 (数量/丰富度/速度 三项) 加权合成"),
    ("排序", "按综合得分降序"),
]
data_hdr = write_meta_block(ws, meta)
ws.append([
    "排名", "区县",
    "中央 (45%)", "行业 (25%)", "市级 (15%)", "区县 (8%)", "公众 (7%)",
    "中央×0.45", "行业×0.25", "市级×0.15", "区县×0.08", "公众×0.07",
    "综合得分 (5 项之和)",
])
for r in ranked:
    name = r["name"]
    c, i, m, d, p = r["central"], r["industry"], r["municipal"], r["district"], r["public"]
    pc, pi, pm, pd, pp = c * 0.45, i * 0.25, m * 0.15, d * 0.08, p * 0.07
    composite = pc + pi + pm + pd + pp
    ws.append([
        r["rank"], name,
        round(c, 2), round(i, 2), round(m, 2), round(d, 2), round(p, 2),
        round(pc, 3), round(pi, 3), round(pm, 3), round(pd, 3), round(pp, 3),
        round(composite, 2),
    ])
style_header_row(ws, data_hdr)
style_data_rows(ws, data_hdr + 1, ws.max_row)
fit_widths(ws, [6, 14] + [12] * 5 + [12] * 5 + [16])

# 保存
Path(OUT_PATH).parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)
print(f"✓ 已写出: {OUT_PATH}")
print(f"  共 {len(wb.sheetnames)} 个 sheet:")
for n in wb.sheetnames:
    print(f"    - {n}")
